from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

from openpyxl import Workbook, load_workbook

from gapsim.emulation.research_registry import DEFAULT_RESEARCH_ROOT
from gapsim.emulation.trench_depo import (
    BOWED_JAR_TRENCH_POINTS,
)

Point = Tuple[float, float]

DEFAULT_STRUCTURE_LIBRARY_PATH = DEFAULT_RESEARCH_ROOT / "structures.xlsx"
DEFAULT_EMULATOR_STRUCTURE_SHEETS = {
    0: "em00_integrated_depo_etch_depth",
}

_INVALID_SHEET_CHARS = re.compile(r"[\[\]\:\*\?\/\\]")


class StructureLibraryError(ValueError):
    pass


def default_emulator_structures() -> Dict[str, List[Point]]:
    return {
        DEFAULT_EMULATOR_STRUCTURE_SHEETS[0]: [(float(x), float(y)) for x, y in BOWED_JAR_TRENCH_POINTS],
    }


def sanitize_structure_name(name: str) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub("_", str(name or "").strip())
    cleaned = re.sub(r"\s+", " ", cleaned).strip("' ")
    if not cleaned:
        cleaned = "structure"
    return cleaned[:31]


def _coerce_points(points: Sequence[Tuple[float, float]]) -> List[Point]:
    coerced = [(float(x), float(y)) for x, y in points]
    if len(coerced) < 2:
        raise StructureLibraryError("A structure needs at least two XY points.")
    return coerced


def _load_or_create_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def list_structure_names(path: Path = DEFAULT_STRUCTURE_LIBRARY_PATH) -> List[str]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        return []
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def read_structure_points(path: Path, sheet_name: str) -> List[Point]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise StructureLibraryError(f"Structure workbook does not exist: {workbook_path}")
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise StructureLibraryError(f"Structure sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        points: List[Point] = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row is None or len(row) < 2:
                continue
            raw_x, raw_y = row[0], row[1]
            if raw_x is None or raw_y is None:
                continue
            try:
                points.append((float(raw_x), float(raw_y)))
            except (TypeError, ValueError):
                continue
        if len(points) < 2:
            raise StructureLibraryError(f"Structure sheet has fewer than two XY points: {sheet_name}")
        return points
    finally:
        wb.close()


def save_structure_points(
    path: Path,
    sheet_name: str,
    points: Sequence[Tuple[float, float]],
) -> str:
    workbook_path = Path(path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    safe_name = sanitize_structure_name(sheet_name)
    coerced = _coerce_points(points)
    wb = _load_or_create_workbook(workbook_path)
    if safe_name in wb.sheetnames:
        index = wb.sheetnames.index(safe_name)
        del wb[safe_name]
        ws = wb.create_sheet(safe_name, index)
    else:
        ws = wb.create_sheet(safe_name)
    ws.append(["x", "y"])
    for x, y in coerced:
        ws.append([float(x), float(y)])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    wb.save(workbook_path)
    wb.close()
    return safe_name


def delete_structure_sheet(path: Path, sheet_name: str) -> str:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise StructureLibraryError(f"Structure workbook does not exist: {workbook_path}")
    safe_name = sanitize_structure_name(sheet_name)
    wb = load_workbook(workbook_path)
    try:
        if safe_name not in wb.sheetnames:
            raise StructureLibraryError(f"Structure sheet not found: {safe_name}")
        del wb[safe_name]
        if wb.sheetnames:
            wb.save(workbook_path)
        else:
            wb.close()
            workbook_path.unlink(missing_ok=True)
            return safe_name
    finally:
        wb.close()
    return safe_name


def ensure_default_structures(
    path: Path = DEFAULT_STRUCTURE_LIBRARY_PATH,
    *,
    overwrite: bool = False,
) -> List[str]:
    workbook_path = Path(path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb = _load_or_create_workbook(workbook_path)
    written: List[str] = []
    try:
        for sheet_name, points in default_emulator_structures().items():
            safe_name = sanitize_structure_name(sheet_name)
            if safe_name in wb.sheetnames and not overwrite:
                continue
            if safe_name in wb.sheetnames:
                index = wb.sheetnames.index(safe_name)
                del wb[safe_name]
                ws = wb.create_sheet(safe_name, index)
            else:
                ws = wb.create_sheet(safe_name)
            ws.append(["x", "y"])
            for x, y in _coerce_points(points):
                ws.append([float(x), float(y)])
            ws.freeze_panes = "A2"
            ws.column_dimensions["A"].width = 16
            ws.column_dimensions["B"].width = 16
            written.append(safe_name)
        wb.save(workbook_path)
    finally:
        wb.close()
    return written
